from PySide6 import QtGui, QtCore, QtWidgets
from ui.ui_interface import Ui_Form
from modules.calibration import Calibration
from modules.manageFiles import ManageFile
from modules.pxrfCalcul import PxrfCalcul
from modules.projectManagment import ProjectManagment
from modules.utils import *
from openpyxl import load_workbook, Workbook
from datetime import datetime, timezone
import os
from math import fabs


class PxrfAnCalculator(Ui_Form, QtWidgets.QWidget, ManageFile, Calibration, PxrfCalcul, ProjectManagment):
    probe_calibration_file, pxrf_calibration_file, unknown_file = None, None, None
    probe_calibration_file_imported, pxrf_calibration_file_imported, unknown_file_imported = False, False, False
    format_probe_reference, format_pxrf_calibration, factor_correction = False, False, False
    do_correction_factor = False
    start_analysis = False
    calibration_chemical_element, non_convertible_oxid = {}, {}
    head_calibration_pxrf_data, calibration_pxrf_data, sample_calibration_frequencies = {}, {}, {}
    correction_factor_for_each_sample, correction_factor = {}, {}
    analysed_unknown_pxrf_data, analysed_unknown_pxrf_data_element = {}, {}
    corrected_data = {}
    ratios, an_content = {}, {}

    def __init__(self):
        super(PxrfAnCalculator, self).__init__()
        # Init Window
        self.setupUi(self)
        self.modification_setup_ui()
        self.setup_connections()
        self.show()

    def modification_setup_ui(self):
        pass

    def setup_connections(self):
        self.pb_browserFile.clicked.connect(self.import_files)
        self.pb_formatProbe.clicked.connect(self.format_probe)
        self.pb_formatPXRF.clicked.connect(self.format_pxrf)
        self.pb_calibrer.clicked.connect(self.correction_factor_calculate)
        self.pb_start.clicked.connect(self.start)
        self.pb_newProject.clicked.connect(self.create_project)
        self.pb_openProject.clicked.connect(self.open_project)
        self.pb_saveProject.clicked.connect(self.save_project)
        self.pb_extractResult.clicked.connect(self.extract_result_xlsx)

    def import_files(self):
        dialog = QtWidgets.QFileDialog()
        dialog.setFileMode(QtWidgets.QFileDialog.AnyFile)
        dialog.setNameFilter('Microsoft Excel Worksheet (*.xlsx)')
        dialog.setViewMode(QtWidgets.QFileDialog.Detail)

        file_names = []
        if dialog.exec():
            file_names = dialog.selectedFiles()[0]

        selected_text = self.select_chooseTypeFile.currentText()

        if str(selected_text) in TYPE_FILE_IMPORTED and TYPE_FILE_IMPORTED[str(selected_text)] == 1:
            if not self.probe_calibration_file_imported and self.probe_calibration_file is None:
                self.probe_calibration_file_imported = True
                self.probe_calibration_file = file_names
                self.cb_probeReference.setCheckable(True)
                self.cb_probeReference.setStyleSheet(u"font-weight: bold;\n"
                                                     "color: rgb(107, 203, 119);")
                self.cb_probeReference.setChecked(True)
                self.cb_probeReference.setEnabled(False)
                self.pb_formatProbe.setEnabled(True)
                if self.pb_formatProbe.isEnabled() and self.pb_formatPXRF.isEnabled():
                    self.pb_calibrer.setEnabled(True)
            else:
                QtWidgets.QMessageBox.warning(self, IMPORTED_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_FILE_IMPORTED)
        elif str(selected_text) in TYPE_FILE_IMPORTED and TYPE_FILE_IMPORTED[str(selected_text)] == 2:
            if not self.pxrf_calibration_file_imported and self.pxrf_calibration_file is None:
                self.pxrf_calibration_file_imported = True
                self.pxrf_calibration_file = file_names
                self.cb_pxrfCalibration.setCheckable(True)
                self.cb_pxrfCalibration.setStyleSheet(u"font-weight: bold;\n"
                                                      "color: rgb(107, 203, 119);")
                self.cb_pxrfCalibration.setChecked(True)
                self.cb_pxrfCalibration.setEnabled(False)
                self.pb_formatPXRF.setEnabled(True)
                if self.pb_formatProbe.isEnabled() and self.pb_formatPXRF.isEnabled():
                    self.pb_calibrer.setEnabled(True)
            else:
                QtWidgets.QMessageBox.warning(self, IMPORTED_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_FILE_IMPORTED)
        elif str(selected_text) in TYPE_FILE_IMPORTED and TYPE_FILE_IMPORTED[str(selected_text)] == 3:
            if not self.unknown_file_imported and self.unknown_file is None:
                self.unknown_file = file_names
                self.cb_unknownAnalysis.setCheckable(True)
                self.cb_unknownAnalysis.setStyleSheet(u"font-weight: bold;\n"
                                                      "color: rgb(107, 203, 119);")
                self.cb_unknownAnalysis.setChecked(True)
                self.cb_unknownAnalysis.setEnabled(False)
            else:
                QtWidgets.QMessageBox.warning(self, IMPORTED_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_FILE_IMPORTED)
        else:
            QtWidgets.QMessageBox.warning(self, FILE_SELECT_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_NOT_SELECT_TYPE_FILE_OR_CHOOSE_IMPORTED_FILE)

    def format_probe(self):
        # REFERENCE PROBE DATA
        if not self.format_probe_reference and len(self.calibration_chemical_element) == 0:
            self.format_probe_reference = True
            vertical_header = []
            workbook, file = self.open_xlsx_file_(self.probe_calibration_file)
            HEAD_PROBE_REFERENCE_DATA_VALUE, PROBE_REFERENCE_DATA = self.compile_xlsx_to_dict(workbook, file)
            self.calibration_chemical_element, self.non_convertible_oxid = self.convert_oxid_to_element(PROBE_REFERENCE_DATA)
            # Tables configurations
            self.table_reference_probe.verticalHeader().setDefaultSectionSize(50)
            self.table_reference_probe.horizontalHeader().setDefaultSectionSize(100)
            self.table_reference_probe.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)

            # Fill Horizontal Header
            HEAD_PROBE_REFERENCE_DATA_VALUE.pop(0)
            for value in HEAD_PROBE_REFERENCE_DATA_VALUE:
                column_count = self.table_reference_probe.columnCount()
                if value in ELEMENTS:
                    item = QtWidgets.QTableWidgetItem(str(ELEMENTS[value].capitalize()))
                else:
                    item = QtWidgets.QTableWidgetItem(str(value).capitalize())
                self.table_reference_probe.insertColumn(column_count)
                self.table_reference_probe.setHorizontalHeaderItem(column_count, item)

            # Fill Vertical Header
            for sample in self.calibration_chemical_element:
                row_count = self.table_reference_probe.rowCount()
                item = QtWidgets.QTableWidgetItem(str(sample))
                self.table_reference_probe.insertRow(row_count)
                self.table_reference_probe.setVerticalHeaderItem(row_count, item)

            # Fill Table
            row_count = self.table_reference_probe.rowCount()
            for sample in self.calibration_chemical_element:
                row = (int(fabs((row_count - self.table_reference_probe.rowCount()))) + 1) - 1
                column_count = self.table_reference_probe.columnCount()
                for element in self.calibration_chemical_element[sample]:
                    item = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.calibration_chemical_element[sample][element])))
                    col = (int(fabs((column_count - (self.table_reference_probe.columnCount())))))
                    self.table_reference_probe.setItem(row, col, item)
                    column_count -= 1
                row_count -= 1

            # Fill Last Column An Content
            row_count = self.table_reference_probe.rowCount()
            for sample in PROBE_REFERENCE_DATA:
                row = (int(fabs((row_count - self.table_reference_probe.rowCount()))) + 1) - 1
                col = self.table_reference_probe.columnCount() - 1
                item = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(PROBE_REFERENCE_DATA[sample][TEXT_CALIBRATION_AN_CONTENT.lower()])))
                self.table_reference_probe.setItem(row, col, item)
                row_count -= 1

            self.pb_formatProbe.setStyleSheet(u'color: rgb(107, 203, 119);')
        else:
            QtWidgets.QMessageBox.warning(self, ALREADY_FORMAT_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_ALREADY_PROBE_FORMAT)

    def format_pxrf(self):
        # CALIBRATION PXRF DATA
        if not self.format_pxrf_calibration and len(self.calibration_pxrf_data) == 0:
            self.format_pxrf_calibration = True
            workbook, file = self.open_xlsx_file_(self.pxrf_calibration_file)
            self.head_calibration_pxrf_data, self.calibration_pxrf_data, self.sample_calibration_frequencies = self.compile_xlsx_calibration_to_dict(workbook, file)
            # Tables configurations
            self.table_calibration_pxrf.verticalHeader().setDefaultSectionSize(50)
            self.table_calibration_pxrf.horizontalHeader().setDefaultSectionSize(100)
            self.table_calibration_pxrf.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)

            # Fill Horizontal Header
            self.head_calibration_pxrf_data.pop(0)
            for value in self.head_calibration_pxrf_data:
                column_count = self.table_calibration_pxrf.columnCount()
                if value in ELEMENTS:
                    item = QtWidgets.QTableWidgetItem(str(value).capitalize())
                else:
                    item = QtWidgets.QTableWidgetItem(str(value).capitalize())
                self.table_calibration_pxrf.insertColumn(column_count)
                self.table_calibration_pxrf.setHorizontalHeaderItem(column_count, item)

            # Fill Vertical Header
            for sample in self.sample_calibration_frequencies:
                for sample_id in range(1, self.sample_calibration_frequencies[sample] + 1):
                    row_count = self.table_calibration_pxrf.rowCount()
                    item = QtWidgets.QTableWidgetItem(str(sample))
                    self.table_calibration_pxrf.insertRow(row_count)
                    self.table_calibration_pxrf.setVerticalHeaderItem(row_count, item)

            # Fill Table
            pass_first_line = False
            row_count = self.table_calibration_pxrf.rowCount()
            for sample in self.calibration_pxrf_data:
                if pass_first_line:
                    row_count += 1
                for sample_id in self.calibration_pxrf_data[sample]:
                    pass_first_line = True
                    row = (int(fabs((row_count - self.table_calibration_pxrf.rowCount()))) + 1) - 1
                    column_count = self.table_calibration_pxrf.columnCount()
                    for element in self.calibration_pxrf_data[sample][sample_id]:
                        item = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.calibration_pxrf_data[sample][sample_id][element])))
                        col = (int(fabs((column_count - (self.table_calibration_pxrf.columnCount())))))
                        self.table_calibration_pxrf.setItem(row, col, item)
                        column_count -= 1
                    row_count -= 1
                row_count -= 1

            self.pb_formatPXRF.setStyleSheet(u'color: rgb(107, 203, 119);')
        else:
            QtWidgets.QMessageBox.warning(self, ALREADY_FORMAT_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_ALREADY_PXRF_FORMAT)

    def correction_factor_calculate(self):
        if self.format_pxrf_calibration and self.format_probe_reference and len(self.calibration_chemical_element) != 0 and len(self.calibration_pxrf_data) != 0:
            if not self.do_correction_factor:
                self.do_correction_factor = True
                self.correction_factor_for_each_sample = self.calcul_correction_factor_for_each_sample(self.calibration_chemical_element, self.calibration_pxrf_data)
                self.correction_factor = Calibration.correction_factor(self.correction_factor_for_each_sample)
                # UNKNOWN DATA
                workbook, file = self.open_xlsx_file_(self.unknown_file)
                HEAD_DATAS_VALUE, self.analysed_unknown_pxrf_data_element = self.compile_xlsx_to_dict(workbook, file)
                unknown_data = self.format_data(self.analysed_unknown_pxrf_data_element)
                self.analysed_unknown_pxrf_data = extract_elements(unknown_data)
                # Correct pxrf unknown data
                self.corrected_data = self.correct_data(self.analysed_unknown_pxrf_data)
                # Tables configurations
                self.table_correction_factor.verticalHeader().setDefaultSectionSize(50)
                self.table_correction_factor.horizontalHeader().setDefaultSectionSize(160)
                self.table_correction_factor.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)
                self.table_pxrf_analysis.verticalHeader().setDefaultSectionSize(50)
                self.table_pxrf_analysis.horizontalHeader().setDefaultSectionSize(160)
                self.table_pxrf_analysis.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)

                # Fill Horizontal Header
                # Table correction factor
                for value in HEAD_CORRECTION_DATA_VALUE:
                    column_count = self.table_correction_factor.columnCount()
                    item = QtWidgets.QTableWidgetItem(str(value).capitalize())
                    self.table_correction_factor.insertColumn(column_count)
                    self.table_correction_factor.setHorizontalHeaderItem(column_count, item)
                # Table pXRF analysis corrected
                for value in HEAD_CORRECTION_DATA_VALUE:
                    column_count = self.table_pxrf_analysis.columnCount()
                    item = QtWidgets.QTableWidgetItem(str(value).capitalize())
                    self.table_pxrf_analysis.insertColumn(column_count)
                    self.table_pxrf_analysis.setHorizontalHeaderItem(column_count, item)

                # Fill Vertical Header
                # Table correction factor
                row_count = self.table_correction_factor.rowCount()
                item = QtWidgets.QTableWidgetItem(str(CORRECTION_FACTOR_TOTAL))
                self.table_correction_factor.insertRow(row_count)
                self.table_correction_factor.setVerticalHeaderItem(row_count, item)
                for sample in self.sample_calibration_frequencies:
                    for sample_id in range(1, self.sample_calibration_frequencies[sample] + 1):
                        row_count = self.table_correction_factor.rowCount()
                        item = QtWidgets.QTableWidgetItem(str(sample))
                        self.table_correction_factor.insertRow(row_count)
                        self.table_correction_factor.setVerticalHeaderItem(row_count, item)
                    for value in VERTICAL_ADDTIONNAL_HEAD_CORRECTION_DATA_VALUE:
                        row_count = self.table_correction_factor.rowCount()
                        item = QtWidgets.QTableWidgetItem(str(value))
                        self.table_correction_factor.insertRow(row_count)
                        self.table_correction_factor.setVerticalHeaderItem(row_count, item)
                # Table pXRF analysis corrected
                for sample in self.analysed_unknown_pxrf_data_element:
                    row_count = self.table_pxrf_analysis.rowCount()
                    item = QtWidgets.QTableWidgetItem(str(sample))
                    self.table_pxrf_analysis.insertRow(row_count)
                    self.table_pxrf_analysis.setVerticalHeaderItem(row_count, item)

                # Fill tables
                # Table correction factor
                # First line on table correction factor
                column_count = self.table_correction_factor.columnCount()
                for element in self.correction_factor:
                    row = 0
                    item = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.correction_factor[element])))
                    col = (int(fabs((column_count - (self.table_correction_factor.columnCount())))))
                    self.table_correction_factor.setItem(row, col, item)
                    column_count -= 1
                # other line on table correction factor
                pass_first_line = False
                row_count = self.table_correction_factor.rowCount() - 1
                for sample in self.correction_factor_for_each_sample:
                    if pass_first_line:
                        row_count += 1
                    for sample_id in self.correction_factor_for_each_sample[sample]:
                        pass_first_line = True
                        row = (int(fabs((row_count - self.table_correction_factor.rowCount()))) + 1) - 1
                        column_count = self.table_correction_factor.columnCount()
                        for element in self.correction_factor_for_each_sample[sample][sample_id]:
                            item = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.correction_factor_for_each_sample[sample][sample_id][element])))
                            col = (int(fabs((column_count - (self.table_correction_factor.columnCount())))))
                            self.table_correction_factor.setItem(row, col, item)
                            column_count -= 1
                        row_count -= 1
                    row_count -= 1

                # Table pXRF analysis corrected
                row_count = self.table_pxrf_analysis.rowCount()
                for sample in self.corrected_data:
                    row = (int(fabs((row_count - self.table_pxrf_analysis.rowCount()))) + 1) - 1
                    column_count = self.table_pxrf_analysis.columnCount()
                    for element in self.corrected_data[sample]:
                        item = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.corrected_data[sample][element])))
                        col = (int(fabs((column_count - (self.table_pxrf_analysis.columnCount())))))
                        self.table_pxrf_analysis.setItem(row, col, item)
                        column_count -= 1
                    row_count -= 1

                # Active start button
                self.pb_start.setEnabled(True)
                self.pb_calibrer.setStyleSheet(u'color: rgb(107, 203, 119);')
        else:
            QtWidgets.QMessageBox.warning(self, PROBE_OR_PXRF_NOT_FORMAT_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_PROBE_OR_PXRF_NOT_FORMAT)

    def start(self):
        if self.do_correction_factor and not self.start_analysis and len(self.corrected_data) != 0:
            self.start_analysis = True
            self.ratios = self.calcul_ratios(self.corrected_data)
            self.an_content = self.calcul_an_content(self.ratios)
            # Tables configurations
            self.table_results.verticalHeader().setDefaultSectionSize(50)
            self.table_results.horizontalHeader().setDefaultSectionSize(150)
            self.table_results.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Fixed)

            # Fill Horizontal Header
            for value in HEAD_ANALYSIS_RESULT:
                column_count = self.table_results.columnCount()
                if value in ELEMENTS:
                    item = QtWidgets.QTableWidgetItem(str(value))
                else:
                    item = QtWidgets.QTableWidgetItem(str(value))
                self.table_results.insertColumn(column_count)
                self.table_results.setHorizontalHeaderItem(column_count, item)

            # Fill Vertical Header
            for sample in self.ratios:
                row_count = self.table_results.rowCount()
                item = QtWidgets.QTableWidgetItem(str(sample))
                self.table_results.insertRow(row_count)
                self.table_results.setVerticalHeaderItem(row_count, item)

            # Fill table
            row_count = self.table_results.rowCount()
            for sample in self.an_content:
                row = (int(fabs((row_count - self.table_results.rowCount()))) + 1) - 1
                # Ratio Ca|Si
                item_ratio_ca_si = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.ratios[sample][TEXT_RATIO_SI])))
                self.table_results.setItem(row, 0, item_ratio_ca_si)
                # An Content Ca|Si
                item_an_content_ca_si = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.an_content[sample][TEXT_AN_CONTENT_RATIO_SI])))
                self.table_results.setItem(row, 1, item_an_content_ca_si)
                # Ratio Ca|Al
                item_ratio_ca_al = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.ratios[sample][TEXT_RATIO_AL])))
                self.table_results.setItem(row, 2, item_ratio_ca_al)
                # An Content Ca|Si
                if self.an_content[sample][TEXT_AN_CONTENT_RATIO_AL] == 0 or self.an_content[sample][TEXT_AN_CONTENT_RATIO_AL] > 99:
                    item_an_content_ca_al = QtWidgets.QTableWidgetItem(str('Non calcul??: Ratio trop grand'))
                else:
                    item_an_content_ca_al = QtWidgets.QTableWidgetItem(str('{:10.3f}'.format(self.an_content[sample][TEXT_AN_CONTENT_RATIO_AL])))
                self.table_results.setItem(row, 3, item_an_content_ca_al)
                row_count -= 1

            self.pb_extractResult.setEnabled(True)

        else:
            QtWidgets.QMessageBox.warning(self, LUNCH_START_VERIFY_ERROR_MSG_BOX_TITLE, ERROR_MESSAGE_LUNCH_START_VERIFY)

    def extract_result_xlsx(self):

        dest_url = QtWidgets.QFileDialog().getExistingDirectory(self, 'Selectionner le dossier', '', QtWidgets.QFileDialog().ShowDirsOnly)

        dest_filename = dest_url + '/' + EXPORTED_FILE_NAME + EXPORTED_FILE_NAME_EXTENSION

        # Remove file if exist in the folder
        if os.path.exists(dest_filename):
            os.remove(dest_filename)

        wb = Workbook()
        # Create all sheet file
        ws_probe_reference_dataset = wb.active
        ws_probe_reference_dataset.title = PROBE_REFRENCE_DATASET_SHEET_NAME
        ws_pxrf_calibration_dataset = wb.create_sheet(title=PXRF_CALIBRATION_DATASET_SHEET_NAME)
        ws_correction_factor_dataset = wb.create_sheet(title=CORRECTION_FACTOR_DATASET_SHEET_NAME)
        ws_corrected_pxrf_dataset = wb.create_sheet(title=CORRECTED_PXRF_DATASET_SHEET_NAME)
        ws_result_dataset = wb.create_sheet(title=RESULTS_DATASET_SHEET_NAME)

        # Table row count
        reference_probe_table_row_count = self.table_reference_probe.rowCount()
        pxrf_calibration_table_row_count = self.table_calibration_pxrf.rowCount()
        correction_factor_table_row_count = self.table_correction_factor.rowCount()
        corrected_pxrf_table_row_count = self.table_pxrf_analysis.rowCount()
        result_an_content_table_row_count = self.table_results.rowCount()
        # Table column count
        reference_probe_table_column_count = self.table_reference_probe.columnCount()
        pxrf_calibration_table_column_count = self.table_calibration_pxrf.columnCount()
        correction_factor_table_column_count = self.table_correction_factor.columnCount()
        corrected_pxrf_table_column_count = self.table_pxrf_analysis.columnCount()
        result_an_content_table_column_count = self.table_results.columnCount()

        # Fill reference probe sheet file
        for row in range(0, reference_probe_table_row_count):
            _ = ws_probe_reference_dataset.cell(column=1, row=(row + 2), value=self.table_reference_probe.verticalHeaderItem(row).text())
        for col in range(0, reference_probe_table_column_count):
            _ = ws_probe_reference_dataset.cell(column=(col + 2), row=1, value=self.table_reference_probe.horizontalHeaderItem(col).text())
        for row in range(0, reference_probe_table_row_count):
            for col in range(0, reference_probe_table_column_count):
                value = self.table_reference_probe.item(row, col).text()
                if value.isdigit():
                    value = float(value)
                _ = ws_probe_reference_dataset.cell(column=(col + 2), row=(row + 2), value=value)

        # Fill pXRF Calibration sheet file
        for row in range(0, pxrf_calibration_table_row_count):
            _ = ws_pxrf_calibration_dataset.cell(column=1, row=(row + 2), value=self.table_calibration_pxrf.verticalHeaderItem(row).text())
        for col in range(0, pxrf_calibration_table_column_count):
            _ = ws_pxrf_calibration_dataset.cell(column=(col + 2), row=1, value=self.table_calibration_pxrf.horizontalHeaderItem(col).text())
        for row in range(0, pxrf_calibration_table_row_count):
            for col in range(0, pxrf_calibration_table_column_count):
                value = self.table_calibration_pxrf.item(row, col).text()
                if value.isdigit():
                    value = float(value)
                _ = ws_pxrf_calibration_dataset.cell(column=(col + 2), row=(row + 2), value=value)

        # Fill Corrector Factor sheet file
        for row in range(0, correction_factor_table_row_count):
            _ = ws_correction_factor_dataset.cell(column=1, row=(row + 2), value=self.table_correction_factor.verticalHeaderItem(row).text())
        for col in range(0, correction_factor_table_column_count):
            _ = ws_correction_factor_dataset.cell(column=(col + 2), row=1, value=self.table_correction_factor.horizontalHeaderItem(col).text())
        for row in range(0, correction_factor_table_row_count):
            for col in range(0, correction_factor_table_column_count):
                value = self.table_correction_factor.item(row, col).text()
                if value.isdigit():
                    value = float(value)
                _ = ws_correction_factor_dataset.cell(column=(col + 2), row=(row + 2), value=value)

        # Fill Corrected pXRF sheet file
        for row in range(0, corrected_pxrf_table_row_count):
            _ = ws_corrected_pxrf_dataset.cell(column=1, row=(row + 2), value=self.table_pxrf_analysis.verticalHeaderItem(row).text())
        for col in range(0, corrected_pxrf_table_column_count):
            _ = ws_corrected_pxrf_dataset.cell(column=(col + 2), row=1, value=self.table_pxrf_analysis.horizontalHeaderItem(col).text())
        for row in range(0, corrected_pxrf_table_row_count):
            for col in range(0, corrected_pxrf_table_column_count):
                value = self.table_pxrf_analysis.item(row, col).text()
                if not value.isdigit():
                    value = float(value)
                _ = ws_corrected_pxrf_dataset.cell(column=(col + 2), row=(row + 2), value=value)

        # Fill result An Content sheet file
        for row in range(0, result_an_content_table_row_count):
            _ = ws_result_dataset.cell(column=1, row=(row + 2), value=self.table_results.verticalHeaderItem(row).text())
        for col in range(0, result_an_content_table_column_count):
            _ = ws_result_dataset.cell(column=(col + 2), row=1, value=self.table_results.horizontalHeaderItem(col).text())
        for row in range(0, result_an_content_table_row_count):
            for col in range(0, result_an_content_table_column_count):
                value = self.table_results.item(row, col).text()
                if value.isdigit():
                    value = float(value)
                _ = ws_result_dataset.cell(column=(col + 2), row=(row + 2), value=value)

        # Save xlsx File
        wb.save(filename=dest_filename)
